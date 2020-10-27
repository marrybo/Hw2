from openpyxl import Workbook

wb = Workbook()

marks = wb.active
marks.title = "Marks"

marks['A1'] = "ФИО"
for i in range(2, 9):
    marks.cell(row=1, column=i).value = ("Рк" + str(i-1))
marks.append(["Петров", 5, 5, 5, 5, 5, 5, 5])
marks.append(["Сидоров", 4, 5, 4, 5, 5, 5, 5])
marks.append(["Иванова", 3, 4, 4, 5, 4, 4, 5])
marks.append(["Кузнецова", 4, 4, 4, 4, 4, 3, 4])
marks.append(["Гаврилов", 5, 4, 4, 3, 4, 4, 5])

ws = wb.create_sheet("INFO")

ws['A1'] = "ФИО"
ws['B1'] = "Сданные профвзносы"
ws['C1'] = "Месяц"
ws['D1'] = "Оценка"
ws.append(["Петров", "+", "сент")
ws.append(["Сидоров", "+", "окт")
ws.append(["Иванова", "+", "сент")
ws.append(["Кузнецова", "-", "окт")
ws.append(["Гаврилов", "-", "сент")

wsh = wb.create_sheet("Result")

wsh['A1'] = "ФИО"
wsh['B1'] = "Летние СЛ"
wsh['C1'] = "Статус"
wsh.append(["Петров", "Охта"])
wsh.append(["Сидоров", "Бауманец"])
wsh.append(["Иванова", "КИТ"])
wsh.append(["Кузнецова", "Бауманец"])
wsh.append(["Гаврилов", "Бухта"])

for i in range(2, 7):
    ws.cell(row=i, column=4).value = round((marks.cell(row=i, column=2).value + marks.cell(row=i, column=3).value + marks.cell(row=i, column=4).value + marks.cell(row=i, column=5).value + marks.cell(row=i, column=6).value + marks.cell(row=i, column=7).value + marks.cell(row=i, column=8).value)/7, 1)

for i in range(2, 7):
    if ws.cell(row=i, column=2).value == "+" and ws.cell(row=i, column=3).value == "сент" and float(ws.cell(row=i, column=4).value) >= 4.5:
        wsh['C' + str(i)] = "+"
    else:
        wsh['C' + str(i)] = "-"

wb.save("Table.xlsx")
